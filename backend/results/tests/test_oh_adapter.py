"""Tests for Ohio SOS results adapter."""
import io

import openpyxl
import pytest

from results.adapters.base import ResultRow
from results.adapters.oh import OhioAdapter, _mark_winners, _parse_master_sheet, _split_name_party


def _build_master_sheet_xlsx(write_in: bool = False) -> bytes:
    """Build a minimal in-memory XLSX matching Ohio's real "Master" sheet
    shape: a two-candidate merged office header, a "Total" row, and a
    trailing per-county row that must be ignored."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master"

    metadata_headers = ["County Name", "Region Name", "Media Market", "Registered Voters", "Ballots Counted", "Official Voter Turnout"]
    for col, header in enumerate(metadata_headers, start=1):
        ws.cell(row=1, column=col, value=None)
        ws.cell(row=2, column=col, value=header)

    ws.cell(row=1, column=7, value="Governor and Lieutenant Governor\n")
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)
    ws.cell(row=2, column=7, value="Jane Smith (D)")
    ws.cell(row=2, column=8, value="John Doe (D)")

    ws.cell(row=1, column=9, value="U.S. Senator\n")
    ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=9)
    candidate9 = "Sam Write (WI)* (D)" if write_in else "Sam Regular (D)"
    ws.cell(row=2, column=9, value=candidate9)

    ws.append(["Total", None, None, None, None, None, 767360, 676562, 214])
    ws.append(["Percentage", None, None, None, None, None, 0.53, 0.47, 1.0])
    ws.append(["Adams", "Southwest", "Cincinnati", 16942, 4070, 0.24, 820, 1905, 100])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestSplitNameParty:
    def test_plain_candidate(self):
        assert _split_name_party("Jon Husted (R)") == ("Jon Husted", "R", False)

    def test_write_in_with_asterisk(self):
        assert _split_name_party("Linda Matthews (WI)* (R)") == ("Linda Matthews", "R", True)

    def test_multi_word_name(self):
        name, party, is_wi = _split_name_party("Casey  Putsch and Kimberly  C. Georgeton (R)")
        assert name == "Casey  Putsch and Kimberly  C. Georgeton"
        assert party == "R"
        assert is_wi is False

    def test_unparseable_string_falls_back_to_raw(self):
        assert _split_name_party("no party marker here") == ("no party marker here", "", False)


class TestParseMasterSheet:
    def test_parses_office_and_vote_counts(self):
        rows = _parse_master_sheet(_build_master_sheet_xlsx())
        by_name = {r.candidate_name: r for r in rows}

        assert by_name["Jane Smith"].vote_count == 767360
        assert by_name["Jane Smith"].office_title == "Governor and Lieutenant Governor"
        assert by_name["John Doe"].vote_count == 676562
        assert by_name["John Doe"].office_title == "Governor and Lieutenant Governor"
        assert by_name["Sam Regular"].vote_count == 214
        assert by_name["Sam Regular"].office_title == "U.S. Senator"

    def test_merged_header_applies_to_every_spanned_column(self):
        rows = _parse_master_sheet(_build_master_sheet_xlsx())
        gov_rows = [r for r in rows if r.office_title == "Governor and Lieutenant Governor"]
        assert len(gov_rows) == 2

    def test_ignores_metadata_and_percentage_and_county_rows(self):
        rows = _parse_master_sheet(_build_master_sheet_xlsx())
        # Only the 3 candidate columns should produce rows, not the county/turnout columns.
        assert len(rows) == 3

    def test_write_in_flagged(self):
        rows = _parse_master_sheet(_build_master_sheet_xlsx(write_in=True))
        write_in_row = next(r for r in rows if r.office_title == "U.S. Senator")
        assert write_in_row.candidate_name == "Sam Write"
        assert write_in_row.is_write_in_aggregate is True

    def test_party_recorded_in_raw(self):
        rows = _parse_master_sheet(_build_master_sheet_xlsx())
        assert all(r.raw["party"] == "D" for r in rows)

    def test_missing_total_row_returns_empty(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Master"
        ws.append(["County Name"])
        ws.append(["Governor"])
        buf = io.BytesIO()
        wb.save(buf)
        assert _parse_master_sheet(buf.getvalue()) == []


class TestMarkWinners:
    def _row(self, office, party, votes, name="X"):
        return ResultRow(
            candidate_name=name, option_label=None, vote_count=votes,
            vote_pct=None, is_winner=None, result_type="official",
            office_title=office, raw={"party": party},
        )

    def test_each_party_gets_its_own_winner(self):
        rows = [
            self._row("Governor", "D", 767360, "Acton"),
            self._row("Governor", "R", 676562, "Ramaswamy"),
            self._row("Governor", "R", 144184, "Putsch"),
        ]
        _mark_winners(rows)
        by_name = {r.candidate_name: r for r in rows}

        assert by_name["Acton"].is_winner is True
        assert by_name["Ramaswamy"].is_winner is True
        assert by_name["Putsch"].is_winner is False

    def test_cross_party_vote_totals_do_not_affect_each_other(self):
        """The Democratic candidate having more raw votes than the Republican
        winner must not suppress the Republican primary winner — they are
        separate contests until the general."""
        rows = [
            self._row("Governor", "D", 1_000_000, "BigD"),
            self._row("Governor", "R", 10, "TinyR"),
        ]
        _mark_winners(rows)
        assert all(r.is_winner for r in rows)

    def test_zero_votes_marks_no_winner(self):
        rows = [self._row("Empty Race", "D", 0, "Nobody")]
        _mark_winners(rows)
        assert rows[0].is_winner is False


class TestOhioAdapter:
    def test_state_is_oh(self):
        assert OhioAdapter.state == "OH"

    def test_returns_none_confidence_when_no_election(self, db):
        from datetime import date

        adapter = OhioAdapter()
        result = adapter.fetch_results(date(2026, 5, 5), 99999)
        assert result.rows == []
        assert result.mapping_confidence == "none"

    def test_returns_none_confidence_when_no_result_files(self, db):
        from datetime import date

        from elections.models import Election

        election = Election.objects.create(
            source_id="oh_sos_2026_primary",
            name="2026 Ohio Primary",
            election_date=date(2026, 5, 5),
            jurisdiction_level="state",
            state="OH",
            status="results_pending",
        )
        adapter = OhioAdapter()
        result = adapter.fetch_results(election.election_date, election.pk)
        assert result.rows == []
        assert result.mapping_confidence == "none"

    def test_mapping_confidence_none_when_all_entries_malformed(self, db):
        from datetime import date

        from elections.models import Election

        election = Election.objects.create(
            source_id="oh_sos_2026_primary_bad",
            name="2026 Ohio Primary",
            election_date=date(2026, 5, 5),
            jurisdiction_level="state",
            state="OH",
            status="results_pending",
            source_metadata={"oh_result_files": [{"party": "REP"}]},  # missing file_url
        )
        adapter = OhioAdapter()
        result = adapter.fetch_results(election.election_date, election.pk)
        assert result.rows == []
        assert result.mapping_confidence == "none"
        assert "missing_file_url" in result.notes


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
