"""
Ohio results adapter.

Ohio publishes official statewide election results as per-party XLSX files
via a JSON index at:
    https://publicfiles.ohiosos.gov/election-results/files-index.json

The index (and every file on publicfiles.ohiosos.gov / data.ohiosos.gov) is
behind Cloudflare Managed Challenge — the aggressive bot-management tier
that fingerprints headless browsers, not just IP reputation. Plain
`requests` and headless Playwright both get 403. Confirmed working: a real,
headed Chrome driven by `nodriver` (undetected-chromedriver's async
successor) against an actual X display. See _fetch_via_browser below.

**Infrastructure**: `_fetch_via_browser` starts its own throwaway `Xvfb`
virtual display for each fetch (via `xvfb` in the Dockerfile) and points
`nodriver` at Playwright's already-downloaded Chromium binary
(`/ms-playwright/.../chrome-linux64/chrome`) — no dependency on a real
desktop session. This is unavoidably heavier than every other adapter
(spins up a real browser process per fetch instead of a plain HTTP
request), which is why Ohio is meant to be triggered manually via the
"Fetch Ohio results now" admin action on Election, not on the regular
scheduled cadence, until proven stable enough for that.

File discovery: fetch files-index.json (via the browser, same CF wall),
find `pastElectionResults[].elections[].fileGroups[].files[]` for the
target year/election, and take the blobPath ending in
`group1/summary-level-official-results-...-{party}.xlsx` (statewide
candidate races by party; there are separate "County Officials Only" and
precinct-level groups not covered here). Full file URL is
`https://publicfiles.ohiosos.gov/election-results/{blobPath}`.
Record the resolved URL(s) in Election.source_metadata["oh_result_files"]:

    Election.source_metadata = {
        "oh_result_files": [
            {"party": "REP", "file_url": "https://publicfiles.ohiosos.gov/election-results/past-elections/2026/.../summary-level-official-results-2026-primary---republican.xlsx"},
            {"party": "DEM", "file_url": "..."},
            {"party": "LIB", "file_url": "..."},
        ]
    }

XLSX shape ("Master" sheet — contains every race; the other 6 sheets are
the same data pre-split by race category, redundant for parsing purposes):
    Row 1: office name, merged across that office's candidate columns
           (e.g. "Governor and Lieutenant Governor" spans 2 columns for a
           2-candidate primary). District/term info is embedded in the
           office name text.
    Row 2: candidate name with party suffix, e.g. "Jon Husted (R)";
           write-in candidates are suffixed "(WI)" before the party.
    Row 3: "Total" — statewide vote count per candidate column.
    Row 4: "Percentage" (not used).
    Row 5+: per-county rows (not used; statewide totals only).
    Columns 1-6 are metadata (County Name, Region Name, Media Market,
    Registered Voters, Ballots Counted, Official Voter Turnout); candidate
    columns start at column 7.
"""
from __future__ import annotations

import asyncio
import base64
import contextlib
import hashlib
import io
import logging
import os
import random
import re
import subprocess
import time

import openpyxl
from django.core.cache import cache

from elections.models import Election

from .base import AdapterResult, ResultRow, StateResultsAdapter
from .registry import register

logger = logging.getLogger(__name__)

_PLAYWRIGHT_BROWSERS_PATH = "/ms-playwright"
_PORTAL_URL = "https://data.ohiosos.gov/portal/past-election-results"
_NAME_PARTY_RE = re.compile(r"^(?P<name>.+?)\s*(?:\((?P<wi>WI)\)\*?\s*)?\((?P<party>[A-Z]+)\)\s*$")
_METADATA_COLUMNS = 6  # County Name, Region Name, Media Market, Registered Voters, Ballots Counted, Turnout


def _resolve_chrome_binary() -> str:
    """Find Playwright's already-downloaded Chromium binary rather than
    hardcoding a revision-specific path (e.g. chromium-1228), which would
    silently break on the next `playwright install` version bump."""
    import glob

    matches = glob.glob(os.path.join(_PLAYWRIGHT_BROWSERS_PATH, "chromium-*", "chrome-linux64", "chrome"))
    if not matches:
        raise RuntimeError(
            f"No Chromium binary found under {_PLAYWRIGHT_BROWSERS_PATH} "
            "(expected Playwright's `playwright install chromium` to have run)"
        )
    return sorted(matches)[-1]


@register
class OhioAdapter(StateResultsAdapter):
    state = "OH"

    VERSION_CACHE_TIMEOUT = 86400 * 30  # 30 days

    @classmethod
    def version_cache_key(cls, election_id: int) -> str:
        return f"oh_sos:hash:{election_id}"

    def fetch_results(self, election_date, election_id: int) -> AdapterResult:
        try:
            election = Election.objects.get(pk=election_id)
        except Election.DoesNotExist:
            logger.error("oh_sos.adapter.missing_election pk=%d", election_id)
            return AdapterResult(
                rows=[], source_url="", mapping_confidence="none",
                notes=f"Election pk={election_id} not found",
            )

        result_files = (election.source_metadata or {}).get("oh_result_files")
        if not result_files:
            logger.warning(
                "oh_sos.adapter.no_result_files election=%s pk=%d",
                election.source_id, election_id,
            )
            return AdapterResult(
                rows=[], source_url="", mapping_confidence="none",
                notes="No oh_result_files in election.source_metadata",
            )

        all_rows: list[ResultRow] = []
        notes_parts: list[str] = []
        fingerprint_parts: list[str] = []

        for entry in result_files:
            file_url = entry.get("file_url")
            party = entry.get("party", "")

            if not file_url:
                logger.warning("oh_sos.adapter.missing_file_url entry=%s", entry)
                notes_parts.append(f"missing_file_url:{party or entry}")
                continue

            try:
                file_bytes = _fetch_via_browser(file_url)
            except Exception as exc:
                logger.warning(
                    "oh_sos.adapter.fetch_error party=%s url=%s err=%s",
                    party, file_url, exc,
                )
                notes_parts.append(f"fetch_error:{party}")
                continue

            fingerprint_parts.append(hashlib.sha256(file_bytes).hexdigest())

            try:
                rows = _parse_master_sheet(file_bytes)
            except Exception as exc:
                logger.warning(
                    "oh_sos.adapter.parse_error party=%s url=%s err=%s",
                    party, file_url, exc,
                )
                notes_parts.append(f"parse_error:{party}")
                continue

            all_rows.extend(rows)

        fingerprint = "|".join(fingerprint_parts)
        cache_key = self.version_cache_key(election_id)
        cached_fingerprint = cache.get(cache_key)
        if fingerprint and fingerprint == cached_fingerprint:
            logger.debug("oh_sos.adapter.unchanged election_id=%d", election_id)
            return AdapterResult(
                rows=[], source_url="", mapping_confidence="full",
                unchanged=True, source_version=fingerprint,
            )

        _mark_winners(all_rows)

        logger.info(
            "oh_sos.adapter.fetched election_id=%d rows=%d files=%d",
            election_id, len(all_rows), len(result_files),
        )

        if not all_rows:
            mapping_confidence = "none"
        elif notes_parts:
            mapping_confidence = "partial"
        else:
            mapping_confidence = "full"

        return AdapterResult(
            rows=all_rows,
            source_url=_PORTAL_URL,
            mapping_confidence=mapping_confidence,
            notes="; ".join(notes_parts),
            source_version=fingerprint,
        )


@contextlib.contextmanager
def _virtual_display():
    """Start a throwaway Xvfb display for the duration of one fetch and set
    DISPLAY to it. A random display number avoids collisions between
    concurrent Celery workers running this adapter at the same time."""
    display_num = random.randint(100, 9999)
    display = f":{display_num}"
    proc = subprocess.Popen(
        ["Xvfb", display, "-screen", "0", "1280x1024x24", "-nolisten", "tcp"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    time.sleep(1)  # give Xvfb a moment to bind before Chrome tries to connect
    old_display = os.environ.get("DISPLAY")
    os.environ["DISPLAY"] = display
    try:
        yield
    finally:
        if old_display is not None:
            os.environ["DISPLAY"] = old_display
        else:
            os.environ.pop("DISPLAY", None)
        proc.terminate()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()


def _fetch_via_browser(file_url: str) -> bytes:
    """Fetch a Cloudflare-walled file via a real headed Chrome session
    running on a throwaway Xvfb virtual display."""
    import nodriver as uc

    async def _fetch() -> bytes:
        browser = await uc.start(
            headless=False,
            browser_executable_path=_resolve_chrome_binary(),
            sandbox=False,  # required under Docker regardless of container user
        )
        try:
            page = await browser.get(_PORTAL_URL)
            await asyncio.sleep(5)
            b64 = await page.evaluate(
                f"""
                fetch({file_url!r})
                    .then(r => r.arrayBuffer())
                    .then(buf => {{
                        let binary = '';
                        const bytes = new Uint8Array(buf);
                        for (let i = 0; i < bytes.byteLength; i++) binary += String.fromCharCode(bytes[i]);
                        return btoa(binary);
                    }})
                """,
                await_promise=True,
                return_by_value=True,
            )
            return base64.b64decode(b64)
        finally:
            try:
                await browser.stop()
            except Exception:
                pass

    with _virtual_display():
        return asyncio.run(_fetch())


def _parse_master_sheet(file_bytes: bytes) -> list[ResultRow]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Master"]

    col_to_office: dict[int, str] = {}
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row != 1:
            continue
        value = ws.cell(row=1, column=merged_range.min_col).value
        if not value:
            continue
        office = value.strip()
        for col in range(merged_range.min_col, merged_range.max_col + 1):
            col_to_office[col] = office

    for col in range(_METADATA_COLUMNS + 1, ws.max_column + 1):
        if col not in col_to_office:
            value = ws.cell(row=1, column=col).value
            if value:
                col_to_office[col] = value.strip()

    total_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        if row[0].value == "Total":
            total_row_idx = row[0].row
            break

    if total_row_idx is None:
        logger.warning("oh_sos.parse.no_total_row")
        return []

    rows: list[ResultRow] = []
    for col, office_title in col_to_office.items():
        if col <= _METADATA_COLUMNS:
            continue

        raw_name = ws.cell(row=2, column=col).value
        if not raw_name:
            continue

        name, party, is_write_in = _split_name_party(str(raw_name))
        votes = _safe_int(ws.cell(row=total_row_idx, column=col).value)

        rows.append(ResultRow(
            candidate_name=name,
            option_label=None,
            vote_count=votes,
            vote_pct=None,
            is_winner=None,  # computed statewide across all party files in _mark_winners
            result_type="official",
            office_title=office_title,
            is_write_in_aggregate=is_write_in,
            raw={"source": "oh_sos_xlsx", "party": party},
        ))

    return rows


def _mark_winners(rows: list[ResultRow]) -> None:
    """Mark the highest vote-getter within each (office, party) group as the
    winner.

    This is a primary: Republican and Democratic candidates for the same
    office are not competing against each other yet, so grouping by
    office_title alone would incorrectly pick a single cross-party
    "winner" and mark the other party's actual primary winner as a loser.
    Grouping by (office_title, party) gives each party's primary its own
    winner, matching reality — both advance to the general.
    """
    by_office_party: dict[tuple[str, str], list[ResultRow]] = {}
    for row in rows:
        party = (row.raw or {}).get("party", "")
        by_office_party.setdefault((row.office_title or "", party), []).append(row)

    for group_rows in by_office_party.values():
        max_votes = max((r.vote_count for r in group_rows), default=0)
        for r in group_rows:
            r.is_winner = r.vote_count == max_votes and max_votes > 0


def _split_name_party(raw: str) -> tuple[str, str, bool]:
    """'Jon Husted (R)' -> ('Jon Husted', 'R', False)
    'Linda Matthews (WI)* (R)' -> ('Linda Matthews', 'R', True)"""
    match = _NAME_PARTY_RE.match(raw.strip())
    if not match:
        return raw.strip(), "", False

    name = match.group("name").strip().rstrip("*").strip()
    party = match.group("party")
    is_write_in = bool(match.group("wi"))
    return name, party, is_write_in


def _safe_int(value) -> int:
    if value is None:
        return 0
    try:
        return int(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0
