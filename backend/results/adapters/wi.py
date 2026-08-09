"""
Wisconsin results adapter — WEC Canvass Reporting System.

Wisconsin has no statewide election-night (unofficial) results feed — 72
independent county clerks, no central live aggregator. See
docs/state-research/WI/WI_Research_v2.md. What the Wisconsin Elections
Commission (WEC) does publish is a single statewide *certified* results
file per election, after county canvass:
    https://elections.wi.gov/elections/election-results

That listing page 403s to automated requests (confirmed via HAR capture in
docs/state-research/WI/), so the direct file URL must be found once (in a
real browser) and recorded in Election.source_metadata["wi_results_xlsx_url"].
The file itself is NOT behind the same block — a plain GET succeeds.

WEC posts the certified file weeks after election day, not on election
night — see the research pass in PR #173 (docs/state-research/WI/WI_Research_v2.md)
for the certified-vs-unofficial split this adapter follows. Election-night
unofficial results are explicitly out of scope (72 independent county
sites, no statewide feed) — tracked as a long-term item on #87/#170.

WEC publishes at least two report shapes, confirmed against two real
captured files (2025 Spring Election "County by County Report" and the
2024 Partisan Primary "Ward by Ward Report" — the latter is what WEC
actually posted for the closest real precedent to a partisan primary; a
"County by County Report" was not published for every contest that cycle).
Both share the same sheet layout, so the parser is written to handle
either without caring which one it was given:
    Sheet 1 ("Document map"): title row, then one contest-title row per
        remaining sheet, in order. Not used for parsing — each contest
        sheet carries its own title, so Sheet 1 is skipped entirely.
    Sheet 2..N (one per contest):
        A few title/blank rows, then a contest title row (e.g.
        "STATE SUPERINTENDENT OF PUBLIC INSTRUCTION", or, for partisan
        contests, "REPRESENTATIVE IN CONGRESS DISTRICT 8 - Republican" —
        confirmed to already match civic_api's "<OFFICE> - <Party>" Stage 1
        title convention, full party name not an abbreviation).
        Then a header row containing a "Total Votes Cast" cell — the
        anchor this parser keys off, rather than a "County" label in
        column A, because the Ward by Ward shape's column-A header cell
        is blank. One party code follows per candidate column ("NP",
        "DEM", "REP", ...).
        Then a candidate-name row, aligned to the header row's columns.
        Then one data row per geography — county name in column A for
        the County by County shape; for Ward by Ward, county name is
        forward-filled only on each county's first ward row and blank on
        every row after, with the municipality/ward label in column B
        instead. Column A therefore can't be used to detect the end of
        data either way — the parser stops when the "Total Votes Cast"
        column stops being numeric, and just sums every data row's
        candidate columns into a statewide total regardless of whether
        the rows are county- or ward-granularity.

A "SCATTERING" column is WEC's catch-all for scattered write-in votes with
no qualified candidate — treated as a write-in aggregate (candidate=None),
same convention as other adapters, and excluded from winner determination.
Named write-in candidates (e.g. "Melby (write-in)") are real Stage-1
candidates and are treated as ordinary candidate rows.

Because this file is WEC's own canvass report, all rows are OFFICIAL —
there is no separate "unofficial election night" version to reconcile.
"""
from __future__ import annotations

import hashlib
import io
import logging

import openpyxl
import requests
from django.core.cache import cache

from elections.models import Election

from .base import AdapterResult, ResultRow, StateResultsAdapter
from .registry import register

logger = logging.getLogger(__name__)

_TIMEOUT = 60


def _safe_int(value) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _find_header_row(ws) -> tuple[int, int] | None:
    """Locate the header row by its "Total Votes Cast" cell rather than a
    "County" label in column A — the County by County shape has one there,
    but the Ward by Ward shape's column-A header cell is blank (county name
    is forward-filled only on each county's first data row, so column A
    can't be used to anchor the header either)."""
    for row in ws.iter_rows(min_row=1, max_row=15):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().lower() == "total votes cast":
                return cell.row, cell.column
    return None


def _parse_contest_sheet(ws) -> list[ResultRow]:
    header_pos = _find_header_row(ws)
    if header_pos is None:
        logger.warning("wi_sos.parse.no_header_row sheet=%s", ws.title)
        return []
    header_row_idx, total_col = header_pos

    office_title = None
    for row in ws.iter_rows(min_row=1, max_row=header_row_idx - 1):
        value = row[0].value
        if isinstance(value, str) and value.strip():
            office_title = value.strip()
    if not office_title:
        logger.warning("wi_sos.parse.no_office_title sheet=%s", ws.title)
        return []

    header_row = ws[header_row_idx]
    name_row = ws[header_row_idx + 1]

    candidate_cols: dict[int, tuple[str, str]] = {}
    for cell in name_row:
        if cell.column <= total_col:
            continue
        name = cell.value.strip() if isinstance(cell.value, str) else None
        if not name:
            continue
        party_cell = header_row[cell.column - 1].value
        party = party_cell.strip() if isinstance(party_cell, str) else ""
        candidate_cols[cell.column] = (name, party)

    if not candidate_cols:
        logger.warning("wi_sos.parse.no_candidate_columns sheet=%s office=%s", ws.title, office_title)
        return []

    # A real data row always has a numeric "Total Votes Cast"; county names
    # are forward-filled in the Ward by Ward shape (blank on every row but
    # each county's first), so column A can't be used to detect the end of
    # data — anchor on the total-votes column going non-numeric instead.
    vote_counts: dict[int, int] = {col: 0 for col in candidate_cols}
    for row in ws.iter_rows(min_row=header_row_idx + 2):
        total_cell = row[total_col - 1].value
        if not isinstance(total_cell, (int, float)):
            break
        for col in candidate_cols:
            vote_counts[col] += _safe_int(row[col - 1].value)

    contest_total = sum(vote_counts.values())
    named_totals = {
        col: votes for col, votes in vote_counts.items()
        if candidate_cols[col][0].strip().upper() != "SCATTERING"
    }
    max_votes = max(named_totals.values(), default=0)

    rows: list[ResultRow] = []
    for col, (name, party) in candidate_cols.items():
        is_scattering = name.strip().upper() == "SCATTERING"
        votes = vote_counts[col]
        rows.append(ResultRow(
            candidate_name=None if is_scattering else name,
            option_label=None,
            vote_count=votes,
            vote_pct=(votes / contest_total * 100) if contest_total else None,
            is_winner=False if is_scattering else (votes == max_votes and max_votes > 0),
            result_type="official",
            office_title=office_title,
            is_write_in_aggregate=is_scattering,
            raw={"source": "wi_wec_xlsx", "party": party},
        ))
    return rows


def _parse_workbook(file_bytes: bytes) -> list[ResultRow]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    rows: list[ResultRow] = []
    for ws in wb.worksheets[1:]:
        rows.extend(_parse_contest_sheet(ws))
    return rows


@register
class WisconsinAdapter(StateResultsAdapter):
    state = "WI"
    VERSION_CACHE_TIMEOUT = 86400 * 30

    @classmethod
    def version_cache_key(cls, election_id: int) -> str:
        return f"wi_wec:hash:{election_id}"

    def fetch_results(self, election_date, election_id: int) -> AdapterResult:
        try:
            election = Election.objects.get(pk=election_id)
        except Election.DoesNotExist:
            logger.error("wi_sos.adapter.missing_election pk=%d", election_id)
            return AdapterResult(
                rows=[], source_url="", mapping_confidence="none",
                notes=f"Election pk={election_id} not found",
            )

        xlsx_url = (election.source_metadata or {}).get("wi_results_xlsx_url")
        if not xlsx_url:
            logger.warning(
                "wi_sos.adapter.no_xlsx_url election=%s pk=%d",
                election.source_id, election_id,
            )
            return AdapterResult(
                rows=[], source_url="", mapping_confidence="none",
                notes="No wi_results_xlsx_url in election.source_metadata",
            )

        try:
            resp = requests.get(xlsx_url, timeout=_TIMEOUT)
            resp.raise_for_status()
        except requests.RequestException as exc:
            logger.warning(
                "wi_sos.adapter.fetch_error url=%s err=%s", xlsx_url, exc,
            )
            return AdapterResult(
                rows=[], source_url=xlsx_url, mapping_confidence="none",
                notes=f"Failed to fetch WEC results file: {exc}",
            )

        fingerprint = hashlib.sha256(resp.content).hexdigest()
        cache_key = self.version_cache_key(election_id)
        if cache.get(cache_key) == fingerprint:
            logger.debug("wi_sos.adapter.unchanged election_id=%d", election_id)
            return AdapterResult(
                rows=[], source_url=xlsx_url, mapping_confidence="full",
                unchanged=True, source_version=fingerprint,
            )

        try:
            rows = _parse_workbook(resp.content)
        except Exception as exc:
            logger.warning("wi_sos.adapter.parse_error url=%s err=%s", xlsx_url, exc)
            return AdapterResult(
                rows=[], source_url=xlsx_url, mapping_confidence="none",
                notes=f"Failed to parse WEC results file: {exc}",
            )

        logger.info(
            "wi_sos.adapter.fetched election_id=%d rows=%d", election_id, len(rows),
        )

        return AdapterResult(
            rows=rows,
            source_url=xlsx_url,
            mapping_confidence="full" if rows else "none",
            source_version=fingerprint,
        )
