"""
Maine results adapter.

Maine has no results API and is not on Clarity Elections. Results are
published as per-race Excel (plurality races) or PDF ranked-choice-voting
summary reports (RCV races) on a JavaScript-rendered page:
    https://www.maine.gov/sos/elections-voting/election-results-data

Because the page is JS-rendered, the per-race file URLs cannot be discovered
automatically at request time — they must be found once (e.g. with a
headless browser) and recorded in Election.source_metadata["me_race_files"]:

    Election.source_metadata = {
        "me_race_files": [
            {
                "office": "US SENATOR",
                "party": "DEM",              # optional, omit for nonpartisan/general
                "district": "",              # e.g. "2" for CD-2, "" for statewide
                "type": "plurality",          # "plurality" (XLSX) or "rcv" (PDF)
                "file_url": "https://www.maine.gov/sos/sites/.../US%20Senate%20DEM%20-%20FINAL.xlsx",
            },
            ...
        ]
    }

Office codes follow Maine's published "Office Abbreviation Key" (reissued
each cycle): US=US Senator, CG=Representative to Congress, GV=Governor,
SS=State Senator, SR=State Representative, JP=Judge of Probate,
RP=Register of Probate, CT=County Treasurer, RD=Register of Deeds,
SH=Sheriff, DA=District Attorney, CC=County Commissioner.

Ranked-choice voting applies to federal offices only (US Senate, US House)
in Maine's general elections; state/county races on the general ballot are
plain plurality. Primaries may use RCV more broadly when a contest has 3+
candidates and no majority winner on first count.

Plurality XLSX shape (one sheet, statewide totals in a "STATE TOTAL" row):
    Row 1: CTY | Municipality | <Candidate 1> | <Candidate 2> | ... | BLANK | TBC
    ...
    Row N: None | STATE TOTAL | <votes 1> | <votes 2> | ... | <blank votes> | <total>

RCV PDF shape (one page, `pdfplumber` extracts as plain text):
    Winner(s) <Name>
    Rounds Round 1 Round 2 ... Round N
    Eliminated <Name> <Name> ...
    Elected <Name>
    <Candidate> <round 1 votes> <round 2 votes> ... <round N votes>
    ...
A candidate's final standing is the last non-zero value in their row —
once eliminated, all later round columns are 0.
"""
from __future__ import annotations

import io
import logging
import re

import openpyxl
import pdfplumber
import requests
from django.core.cache import cache

from elections.models import Election

from .base import AdapterResult, ResultRow, StateResultsAdapter
from .registry import register

logger = logging.getLogger(__name__)

_TIMEOUT = 30
_NON_CANDIDATE_COLUMNS = {"blank", "tbc", "void", "scattering"}
_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; CivicMirror/1.0; +https://civicmirror.app)"
    ),
}


@register
class MaineAdapter(StateResultsAdapter):
    state = "ME"

    VERSION_CACHE_TIMEOUT = 86400 * 30  # 30 days

    @classmethod
    def version_cache_key(cls, election_id: int) -> str:
        return f"me_sos:hash:{election_id}"

    def fetch_results(self, election_date, election_id: int) -> AdapterResult:
        try:
            election = Election.objects.get(pk=election_id)
        except Election.DoesNotExist:
            logger.error("me_sos.adapter.missing_election pk=%d", election_id)
            return AdapterResult(
                rows=[], source_url="", mapping_confidence="none",
                notes=f"Election pk={election_id} not found",
            )

        race_files = (election.source_metadata or {}).get("me_race_files")
        if not race_files:
            logger.warning(
                "me_sos.adapter.no_race_files election=%s pk=%d",
                election.source_id, election_id,
            )
            return AdapterResult(
                rows=[], source_url="", mapping_confidence="none",
                notes="No me_race_files in election.source_metadata",
            )

        all_rows: list[ResultRow] = []
        notes_parts: list[str] = []
        fingerprint_parts: list[str] = []

        for entry in race_files:
            file_url = entry.get("file_url")
            office = entry.get("office", "")
            party = entry.get("party", "")
            district = entry.get("district", "")
            race_type = entry.get("type", "plurality")

            if not file_url:
                logger.warning("me_sos.adapter.missing_file_url entry=%s", entry)
                continue

            try:
                resp = requests.get(file_url, headers=_HEADERS, timeout=_TIMEOUT)
                resp.raise_for_status()
            except requests.RequestException as exc:
                logger.warning(
                    "me_sos.adapter.fetch_error office=%s url=%s err=%s",
                    office, file_url, exc,
                )
                notes_parts.append(f"fetch_error:{office}")
                continue

            fingerprint_parts.append(resp.content.hex()[:16])

            office_title = _build_office_title(office, party, district)

            try:
                if race_type == "rcv":
                    rows = _parse_rcv_pdf(resp.content, office_title)
                else:
                    rows = _parse_plurality_xlsx(resp.content, office_title)
            except Exception as exc:
                logger.warning(
                    "me_sos.adapter.parse_error office=%s url=%s err=%s",
                    office, file_url, exc,
                )
                notes_parts.append(f"parse_error:{office}")
                continue

            all_rows.extend(rows)

        fingerprint = "".join(fingerprint_parts)
        cache_key = self.version_cache_key(election_id)
        cached_fingerprint = cache.get(cache_key)
        if fingerprint and fingerprint == cached_fingerprint:
            logger.debug("me_sos.adapter.unchanged election_id=%d", election_id)
            return AdapterResult(
                rows=[], source_url="", mapping_confidence="full",
                unchanged=True, source_version=fingerprint,
            )

        logger.info(
            "me_sos.adapter.fetched election_id=%d rows=%d races=%d",
            election_id, len(all_rows), len(race_files),
        )

        return AdapterResult(
            rows=all_rows,
            source_url="https://www.maine.gov/sos/elections-voting/election-results-data",
            mapping_confidence="full" if not notes_parts else "partial",
            notes="; ".join(notes_parts),
            source_version=fingerprint,
        )


def _build_office_title(office: str, party: str, district: str) -> str:
    title = office
    if district:
        title = f"{title} District {district}"
    if party:
        title = f"{title} - {party}"
    return title


def _parse_plurality_xlsx(file_bytes: bytes, office_title: str) -> list[ResultRow]:
    """Parse a Maine SOS plurality-race XLSX: candidate columns, town rows,
    ending in a 'STATE TOTAL' row with statewide vote counts."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [c.value for c in row]
        if any(isinstance(v, str) and v.strip().upper() in ("CTY", "OFFICE") for v in values):
            header_row = row
            break

    if header_row is None:
        logger.warning("me_sos.parse.no_header_row office=%s", office_title)
        return []

    candidate_cols: dict[int, str] = {}
    for cell in header_row:
        value = cell.value.strip() if isinstance(cell.value, str) else None
        if not value:
            continue
        if value.strip().lower() in _NON_CANDIDATE_COLUMNS:
            continue
        if value.upper() in ("CTY", "MUNICIPALITY", "OFFICE", "TOWN", "WARD-PRECINCT"):
            continue
        candidate_cols[cell.column] = value

    if not candidate_cols:
        logger.warning("me_sos.parse.no_candidate_columns office=%s", office_title)
        return []

    state_total_row = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "STATE TOTAL" in cell.value.upper():
                state_total_row = row
                break
        if state_total_row:
            break

    if state_total_row is None:
        logger.warning("me_sos.parse.no_state_total_row office=%s", office_title)
        return []

    vote_counts: dict[str, int] = {}
    for col, candidate_name in candidate_cols.items():
        cell = state_total_row[col - 1]
        vote_counts[candidate_name] = _safe_int(cell.value)

    max_votes = max(vote_counts.values(), default=0)
    total_votes = sum(vote_counts.values())

    rows: list[ResultRow] = []
    for candidate_name, votes in vote_counts.items():
        rows.append(ResultRow(
            candidate_name=candidate_name,
            option_label=None,
            vote_count=votes,
            vote_pct=(votes / total_votes * 100) if total_votes else None,
            is_winner=(votes == max_votes and max_votes > 0),
            result_type="official",
            office_title=office_title,
            raw={"source": "me_sos_xlsx"},
        ))
    return rows


_ROUND_LINE_RE = re.compile(r"^Rounds\s+(.*)$", re.MULTILINE)
_WINNER_LINE_RE = re.compile(r"^Winner\(s\)\s+(.*)$", re.MULTILINE)
_CANDIDATE_ROW_RE = re.compile(r"^([A-Za-z][A-Za-z ,.'\-]*?)\s+((?:\d[\d,]*\s*)+)$", re.MULTILINE)


def _parse_rcv_pdf(file_bytes: bytes, office_title: str) -> list[ResultRow]:
    """Parse a Maine SOS RCV Summary Report PDF: single page, plain text,
    with a Winner(s) line and a per-candidate row of round-by-round vote
    counts (0 in all columns after the round a candidate was eliminated)."""
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    winner_match = _WINNER_LINE_RE.search(text)
    winner_name = winner_match.group(1).strip() if winner_match else None

    rounds_match = _ROUND_LINE_RE.search(text)
    if not rounds_match:
        logger.warning("me_sos.parse.no_rounds_line office=%s", office_title)
        return []
    num_rounds = len(re.findall(r"Round\s+\d+", rounds_match.group(1)))
    if num_rounds == 0:
        logger.warning("me_sos.parse.no_round_count office=%s", office_title)
        return []

    skip_labels = {
        "eliminated", "elected", "exhausted ballots", "threshold",
        "contest", "jurisdiction", "office", "date", "rounds",
    }

    rows: list[ResultRow] = []
    for match in _CANDIDATE_ROW_RE.finditer(text):
        name = match.group(1).strip()
        if name.lower() in skip_labels:
            continue

        numbers = [_safe_int(v) for v in match.group(2).split()]
        if len(numbers) != num_rounds:
            continue

        # Final standing = last non-zero round value (0 after elimination).
        final_votes = 0
        for v in numbers:
            if v > 0:
                final_votes = v

        rows.append(ResultRow(
            candidate_name=name,
            option_label=None,
            vote_count=final_votes,
            vote_pct=None,
            is_winner=(winner_name is not None and _names_match(name, winner_name)),
            result_type="official",
            office_title=office_title,
            raw={"source": "me_sos_rcv_pdf", "rounds": numbers},
        ))
    return rows


def _names_match(a: str, b: str) -> bool:
    def normalize(s: str) -> str:
        return re.sub(r"[^a-z]", "", s.lower())
    return normalize(a) == normalize(b)


def _safe_int(value) -> int:
    if value is None:
        return 0
    try:
        return int(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0
