import datetime as dt
import io
from pathlib import Path

from openpyxl import load_workbook

from integrations.al_sos.parsers import (
    normalize_contest_title,
    parse_election_year_page,
    parse_enr_workbook,
    parse_fcpa_committee_detail,
    parse_fcpa_race_search_response,
)

FIXTURES = Path(__file__).parent / "fixtures"


def test_normalize_contest_title_strips_party_suffix():
    assert normalize_contest_title("LIEUTENANT GOVERNOR (REP)") == ("LIEUTENANT GOVERNOR", "REP")
    assert normalize_contest_title("STATE REPRESENTATIVE, DISTRICT 63") == (
        "STATE REPRESENTATIVE, DISTRICT 63",
        "",
    )


def test_parse_enr_workbook_aggregates_county_rows():
    content = (FIXTURES / "al_sos_enr_export.xlsx").read_bytes()

    parsed = parse_enr_workbook(content)

    lieutenant_governor = [
        row
        for row in parsed.rows
        if row.office_title == "LIEUTENANT GOVERNOR" and row.raw["party_code"] == "REP"
    ]
    by_candidate = {row.candidate_name: row for row in lieutenant_governor}

    assert "Wes Allen" in by_candidate
    assert "John Wahl" in by_candidate
    assert by_candidate["Wes Allen"].vote_count > 13036
    assert by_candidate["John Wahl"].vote_count > 15588
    assert all(row.result_type == "official" for row in parsed.rows)
    assert parsed.is_complete is True
    assert parsed.source_version.startswith("1001295:")
    assert parsed.county_stats["01"]["precincts_reported"] == 177


def test_parse_enr_workbook_preserves_party_from_column_when_title_has_suffix():
    content = (FIXTURES / "al_sos_enr_export.xlsx").read_bytes()

    parsed = parse_enr_workbook(content)
    row = next(row for row in parsed.rows if row.candidate_name == "Wes Allen")

    assert row.raw["party_code"] == "REP"
    assert row.raw["contest_code"] == "00100892"
    assert row.raw["source"] == "al_sos_enr"


def test_source_version_changes_when_raw_workbook_row_counts_change():
    content = (FIXTURES / "al_sos_enr_export.xlsx").read_bytes()
    parsed = parse_enr_workbook(content)

    all_results_extra = _with_appended_row(content, "AllResults", row_number=2)
    statistics_extra = _with_appended_row(content, "Statistics", row_number=2)

    assert parse_enr_workbook(all_results_extra).source_version != parsed.source_version
    assert parse_enr_workbook(statistics_extra).source_version != parsed.source_version


def test_parse_enr_workbook_treats_overreported_precincts_as_complete():
    content = (FIXTURES / "al_sos_enr_export.xlsx").read_bytes()
    workbook = load_workbook(io.BytesIO(content))
    statistics = workbook["Statistics"]
    headers = [cell.value for cell in statistics[1]]
    precincts_reported_column = headers.index("Precincts Reported") + 1
    total_precincts_column = headers.index("Total Precincts") + 1
    statistics.cell(row=2, column=precincts_reported_column).value = (
        statistics.cell(row=2, column=total_precincts_column).value + 1
    )

    parsed = parse_enr_workbook(_save_workbook(workbook))

    assert parsed.is_complete is True
    assert all(row.result_type == "official" for row in parsed.rows)


def _with_appended_row(content: bytes, sheet_name: str, row_number: int) -> bytes:
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook[sheet_name]
    sheet.append([cell.value for cell in sheet[row_number]])
    return _save_workbook(workbook)


def _save_workbook(workbook) -> bytes:
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _year_page_html() -> str:
    return (FIXTURES / "al_year_page_2026.html").read_text()


def test_parse_election_year_page_finds_all_headings():
    elections = parse_election_year_page(_year_page_html())

    assert len(elections) == 4
    names = [e["name"] for e in elections]
    assert "Special General Election House District 63" in names
    assert "Primary Election" in names
    assert "Primary Runoff Election" in names
    assert "General Election" in names


def test_parse_election_year_page_extracts_dates_and_types():
    elections = parse_election_year_page(_year_page_html())
    by_name = {e["name"]: e for e in elections}

    assert by_name["Primary Election"]["election_date"] == dt.date(2026, 5, 19)
    assert by_name["Primary Election"]["election_type"] == "primary"
    assert by_name["Primary Runoff Election"]["election_date"] == dt.date(2026, 6, 16)
    assert by_name["Primary Runoff Election"]["election_type"] == "primary_runoff"
    assert by_name["General Election"]["election_date"] == dt.date(2026, 11, 3)
    assert by_name["General Election"]["election_type"] == "general"
    assert by_name["Special General Election House District 63"]["election_type"] == "special"


def test_parse_election_year_page_extracts_document_links():
    elections = parse_election_year_page(_year_page_html())
    by_name = {e["name"]: e for e in elections}

    primary_links = by_name["Primary Election"]["document_links"]
    assert {"label": "Sample Ballots", "url": "https://www.sos.alabama.gov/alabama-votes/2026-primary-election-sample-ballots"} in primary_links
    republican_cert = next(link for link in primary_links if "Republican Party Certification" in link["label"])
    assert republican_cert["url"] == "https://www.sos.alabama.gov/sites/default/files/election-2026/2026RepublicanCertification.pdf"


def test_parse_election_year_page_source_id_is_stable_and_unique():
    elections = parse_election_year_page(_year_page_html())
    source_ids = [e["source_id"] for e in elections]

    assert len(source_ids) == len(set(source_ids))
    assert all(sid.startswith("al_sos_2026_") for sid in source_ids)


def test_parse_election_year_page_does_not_split_on_plain_hyphen_in_name():
    html = (
        "<h3>City Council District 63 - Runoff – January 13, 2026</h3>"
        "<blockquote><p><a href=\"/x.pdf\">Doc</a></p></blockquote>"
    )
    elections = parse_election_year_page(html)

    assert len(elections) == 1
    assert elections[0]["name"] == "City Council District 63 - Runoff"
    assert elections[0]["election_date"] == dt.date(2026, 1, 13)


def test_parse_election_year_page_classifies_general_runoff():
    html = "<h3>General Election Runoff – December 1, 2026</h3><blockquote></blockquote>"
    elections = parse_election_year_page(html)

    assert elections[0]["election_type"] == "general_runoff"


def _search_results_json() -> str:
    return (FIXTURES / "al_fcpa_search_results_page1.json").read_text()


def _committee_detail_html() -> str:
    return (FIXTURES / "al_fcpa_committee_detail_4834.html").read_text()


def test_parse_fcpa_race_search_response_returns_rows_and_total():
    rows, total_records = parse_fcpa_race_search_response(_search_results_json())

    assert total_records == 848
    assert len(rows) == 3
    assert rows[0]["committee_id"] == 4834
    assert rows[0]["candidate_name"] == "ABBETT, JIMMY"
    assert rows[0]["candidate_status"] == "Active"
    assert rows[0]["year"] == 2026


def test_parse_fcpa_committee_detail_extracts_structured_fields():
    detail = parse_fcpa_committee_detail(_committee_detail_html())

    assert detail["committee_id"] == 4834
    assert detail["candidateFirstName"] == "JIMMY"
    assert detail["candidateLastName"] == "ABBETT"
    assert detail["suffix"] == ""
    assert detail["office"] == "Sheriff"
    assert detail["jurisdiction"] == "TALLAPOOSA COUNTY"
    assert detail["district"] == ""
    assert detail["party"] == "Republican"
    assert detail["committeeStatus"] == "Active"
    assert detail["dissolved"] is False
