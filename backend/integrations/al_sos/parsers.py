from __future__ import annotations

import datetime as dt
import io
import json
import re as _re
from collections import defaultdict
from dataclasses import dataclass
from urllib.parse import urljoin

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from results.adapters.base import ResultRow

from .exceptions import AlSosError

_PARTY_SUFFIX_RE = _re.compile(r"\s+\(([A-Z]{2,5})\)\s*$")

_YEAR_PAGE_BASE_URL = "https://www.sos.alabama.gov"
_DASH_RE = _re.compile(r"\s–\s")  # en dash only -- do not add ASCII hyphen-minus here,
# real election names may contain a plain " - " (e.g. "District 63 - Runoff") and a
# hyphen-inclusive class would split on the wrong dash and silently drop the election.


@dataclass(frozen=True)
class AlEnrParsedResult:
    rows: list[ResultRow]
    source_version: str
    is_complete: bool
    county_stats: dict[str, dict]


def normalize_contest_title(title: str) -> tuple[str, str]:
    normalized = " ".join(str(title or "").split())
    match = _PARTY_SUFFIX_RE.search(normalized)
    if not match:
        return normalized, ""
    return _PARTY_SUFFIX_RE.sub("", normalized).strip(), match.group(1)


def parse_enr_workbook(content: bytes) -> AlEnrParsedResult:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if "AllResults" not in workbook.sheetnames:
        raise AlSosError("Alabama ENR workbook missing AllResults sheet")
    if "Statistics" not in workbook.sheetnames:
        raise AlSosError("Alabama ENR workbook missing Statistics sheet")

    county_stats, statistics_row_count = _parse_statistics(workbook["Statistics"])
    is_complete = all(
        stat["precincts_reported"] >= stat["total_precincts"]
        for stat in county_stats.values()
        if stat["total_precincts"] is not None
    )
    result_type = "official" if is_complete else "unofficial"

    totals: dict[tuple[str, str, str, str], int] = defaultdict(int)
    metadata: dict[tuple[str, str, str, str], dict] = {}
    election_codes: set[str] = set()
    all_results_row_count = 0

    for raw in _iter_dict_rows(workbook["AllResults"]):
        all_results_row_count += 1
        contest_code = _clean(raw.get("Contest Code"))
        contest_title = _clean(raw.get("Contest Title"))
        candidate_name = _clean(raw.get("Candidate Name"))
        party_code = _clean(raw.get("Party Code"))
        votes = _safe_int(raw.get("Votes"))
        election_code = _clean(raw.get("Election Code"))
        county_code = _clean(raw.get("County Code"))
        if not contest_code or not contest_title or not candidate_name:
            continue

        office_title, party_from_title = normalize_contest_title(contest_title)
        party = party_code or party_from_title
        key = (contest_code, office_title, candidate_name, party)
        totals[key] += votes
        election_codes.add(election_code)
        metadata.setdefault(
            key,
            {
                "contest_code": contest_code,
                "contest_title": contest_title,
                "party_code": party,
                "source": "al_sos_enr",
                "county_codes": [],
            },
        )
        metadata[key]["county_codes"].append(county_code)

    rows = [
        ResultRow(
            office_title=office_title,
            candidate_name=None if _is_write_in(candidate_name) else candidate_name,
            option_label=None,
            vote_count=vote_count,
            vote_pct=None,
            is_winner=None,
            result_type=result_type,
            is_write_in_aggregate=_is_write_in(candidate_name),
            raw=metadata[key],
        )
        for key, vote_count in sorted(totals.items(), key=lambda item: item[0])
        for _contest_code, office_title, candidate_name, _party in [key]
    ]

    return AlEnrParsedResult(
        rows=rows,
        source_version=_source_version(
            election_codes,
            county_stats,
            all_results_row_count=all_results_row_count,
            statistics_row_count=statistics_row_count,
        ),
        is_complete=is_complete,
        county_stats=county_stats,
    )


def _parse_statistics(sheet) -> tuple[dict[str, dict], int]:
    stats: dict[str, dict] = {}
    row_count = 0
    for raw in _iter_dict_rows(sheet):
        row_count += 1
        county_code = _clean(raw.get("County Code"))
        if not county_code:
            continue
        last_updated = raw.get("Last Updated")
        if isinstance(last_updated, dt.datetime):
            last_updated_value = last_updated.isoformat()
        else:
            last_updated_value = _clean(last_updated)
        stats[county_code] = {
            "ballots_cast": _safe_int(raw.get("Ballots Cast")),
            "total_precincts": _safe_int(raw.get("Total Precincts")),
            "precincts_reported": _safe_int(raw.get("Precincts Reported")),
            "last_updated": last_updated_value,
        }
    return stats, row_count


def _iter_dict_rows(sheet):
    rows = sheet.iter_rows(values_only=True)
    headers = [_clean(value) for value in next(rows, [])]
    for row in rows:
        yield dict(zip(headers, row))


def _source_version(
    election_codes: set[str],
    county_stats: dict[str, dict],
    *,
    all_results_row_count: int,
    statistics_row_count: int,
) -> str:
    code = ",".join(sorted(election_codes))
    latest = max((stat["last_updated"] for stat in county_stats.values()), default="")
    return f"{code}:{latest}:all_results={all_results_row_count}:statistics={statistics_row_count}"


def _safe_int(value) -> int:
    if value in (None, ""):
        return 0
    return int(str(value).replace(",", "").strip())


def _clean(value) -> str:
    return " ".join(str(value or "").split())


def _is_write_in(value: str) -> bool:
    return "write-in" in value.lower() or "write in" in value.lower()


def _slugify(text: str) -> str:
    return _re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def _infer_election_type(heading_text: str) -> str:
    lowered = heading_text.lower()
    if "special" in lowered:
        return "special"
    if "runoff" in lowered and "general" in lowered:
        return "general_runoff"
    if "runoff" in lowered:
        return "primary_runoff"
    if "primary" in lowered:
        return "primary"
    if "general" in lowered:
        return "general"
    if "municipal" in lowered:
        return "municipal"
    return "other"


def parse_election_year_page(html: str) -> list[dict]:
    """
    Parse an Alabama SOS year-specific Election Information page
    (www.sos.alabama.gov/alabama-votes/voter/election-information/{year}).

    Each <h3> heading is "{Name} – {Month Day, Year}"; the immediately
    following <blockquote> holds that election's official document links.
    """
    soup = BeautifulSoup(html, "html.parser")
    results: list[dict] = []

    for heading in soup.find_all("h3"):
        text = " ".join(heading.get_text().split())
        parts = _DASH_RE.split(text, maxsplit=1)
        if len(parts) != 2:
            continue
        name, date_text = parts[0].strip(), parts[1].strip()
        try:
            election_date = dt.datetime.strptime(date_text, "%B %d, %Y").date()
        except ValueError:
            continue

        blockquote = heading.find_next_sibling("blockquote")
        document_links = []
        if blockquote is not None:
            for a in blockquote.find_all("a", href=True):
                label = " ".join(a.get_text().split())
                url = urljoin(_YEAR_PAGE_BASE_URL, a["href"])
                document_links.append({"label": label, "url": url})

        results.append({
            "name": name,
            "election_date": election_date,
            "election_type": _infer_election_type(name),
            "source_id": f"al_sos_{election_date.year}_{_slugify(name)}",
            "document_links": document_links,
        })

    return results


def parse_fcpa_race_search_response(json_text: str) -> tuple[list[dict], int]:
    """Parse a com.acf.common.page.politicalracesearchresults JSON response."""
    payload = json.loads(json_text)
    if not payload.get("success"):
        raise AlSosError("Alabama FCPA race search response reported success=false")

    data = payload.get("data") or {}
    rows = [
        {
            "committee_id": row["COMMITTEEID"],
            "candidate_name": _clean(row.get("CANDIDATE", "")),
            "candidate_status": row.get("CANDIDATESTATUS", ""),
            "year": row.get("YEAR"),
        }
        for row in data.get("list", [])
    ]
    return rows, int(data.get("totalRecords", 0))


def _extract_balanced_object(text: str, marker: str) -> str:
    """Extract the {...} object literal immediately following `marker`."""
    start = text.find(marker)
    if start == -1:
        raise AlSosError(f"Alabama FCPA committee detail page missing {marker!r}")
    brace_start = text.find("{", start)
    if brace_start == -1:
        raise AlSosError("Alabama FCPA committee detail page missing object literal")

    depth = 0
    for i in range(brace_start, len(text)):
        if text[i] == "{":
            depth += 1
        elif text[i] == "}":
            depth -= 1
            if depth == 0:
                return text[brace_start:i + 1]
    raise AlSosError("Alabama FCPA committee detail page has unterminated object literal")


def parse_fcpa_committee_detail(html: str) -> dict:
    """
    Parse the committeeDetailsObj JSON embedded in a committee detail page
    (page.acfPublicCommitteeDetails). Verified strict JSON against the real
    capture in docs/state-research/AL/fcpa.alabamavotes.gov_Archive
    [26-07-20 12-42-17].har -- json.loads works directly, no JS-literal
    normalization needed.
    """
    raw = _extract_balanced_object(html, "committeeDetailsObj")
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise AlSosError(f"Alabama FCPA committeeDetailsObj is not valid JSON: {exc}") from exc

    return {
        "committee_id": data.get("id"),
        "candidateFirstName": data.get("candidateFirstName", ""),
        "candidateMiddleName": data.get("candidateMiddleName", ""),
        "candidateLastName": data.get("candidateLastName", ""),
        "suffix": data.get("suffix", ""),
        "office": data.get("office", ""),
        "jurisdiction": data.get("jurisdiction", ""),
        "district": data.get("district", ""),
        "place": data.get("place", ""),
        "party": data.get("party", ""),
        "committeeStatus": data.get("committeeStatus", ""),
        "dissolved": bool(data.get("dissolved", False)),
    }
