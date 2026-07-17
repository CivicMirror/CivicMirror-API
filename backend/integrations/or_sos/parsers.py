from __future__ import annotations

import csv
import datetime
import hashlib
import io
import re
import zipfile
from dataclasses import dataclass
from pathlib import PurePosixPath

import pdfplumber
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .exceptions import OrSosUnsupportedDocumentError


@dataclass(frozen=True)
class OrResultRecord:
    office_title: str
    choice: str
    vote_count: int
    vote_pct: float | None = None
    party: str = ""
    jurisdiction: str = ""
    source_file: str = ""


@dataclass(frozen=True)
class OrOpenOffice:
    office_title: str
    office_code: str
    district: str = ""
    ocd_division_id: str = "ocd-division/country:us/state:or"
    geography_scope: str = "statewide"


@dataclass(frozen=True)
class OrElectionInfo:
    name: str
    election_date: str
    election_type: str
    source_url: str = ""


@dataclass(frozen=True)
class OrBallotReturn:
    election: str
    election_date: datetime.date
    count_date: datetime.date
    daily_ballots_returned: int
    cumulative_ballots_returned: int
    daily_return_pct_of_total_ballots: float | None = None
    daily_return_pct_of_total_return: float | None = None
    cumulative_return_pct_of_total_ballots: float | None = None
    cumulative_return_pct_of_total_return: float | None = None


@dataclass(frozen=True)
class OrCandidateFiling:
    ballot_name: str
    party: str
    office: str
    election: str
    filing_method: str
    filing_date: str
    qualified: str


@dataclass(frozen=True)
class OrLocalMeasure:
    measure_number: str
    election: str
    county: str
    ballot_title_caption: str


def document_checksum(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def parse_result_document(content: bytes, source_url: str) -> list[OrResultRecord]:
    suffix = PurePosixPath(source_url.split("?", 1)[0]).suffix.lower()
    if content.startswith(b"%PDF") or suffix == ".pdf":
        return _parse_pdf(content)
    if suffix == ".zip":
        return _parse_zip(content)
    if suffix in {".csv", ".txt", ".tsv"}:
        return _parse_delimited(content, source_file=PurePosixPath(source_url).name)
    if suffix == ".xlsx":
        return _parse_xlsx(content, source_file=PurePosixPath(source_url).name)
    if suffix == ".xls":
        raise OrSosUnsupportedDocumentError(f"Oregon result document type is not parsed yet: {suffix}")
    raise OrSosUnsupportedDocumentError(f"Unsupported Oregon result document type: {suffix or 'unknown'}")


def _parse_pdf(content: bytes) -> list[OrResultRecord]:
    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception as exc:
        raise OrSosUnsupportedDocumentError(f"Oregon result PDF could not be parsed: {exc}") from exc
    return parse_result_pdf_text(text)


def parse_result_pdf_text(text: str) -> list[OrResultRecord]:
    records: list[OrResultRecord] = []
    page_chunks = re.split(r"(?=May\s+\d{1,2},\s+\d{4},\s+Primary Election Abstract of Votes)", text)
    for chunk in page_chunks:
        records.extend(_parse_result_pdf_page(chunk))
    return records


def parse_open_offices_pdf(content: bytes) -> list[OrOpenOffice]:
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    return parse_open_offices_text(text)


def parse_election_page(html: str, source_url: str = "") -> list[OrElectionInfo]:
    soup = BeautifulSoup(html, "lxml")
    text = soup.get_text("\n")
    return parse_election_text(text, source_url=source_url)


def parse_election_text(text: str, source_url: str = "") -> list[OrElectionInfo]:
    elections: list[OrElectionInfo] = []
    seen: set[tuple[str, str]] = set()
    lines = [_clean_text(line) for line in text.splitlines()]
    pending_type = ""

    for line in lines:
        election = _election_from_line(line, source_url)
        if not election:
            pending_type = _heading_election_type_from_line(line) or pending_type
            date_match = _DATE_RE.search(line)
            if pending_type and date_match:
                election = _election_info_from_date(date_match.group(0), pending_type, source_url)
        if not election:
            continue
        key = (election.election_date, election.election_type)
        if key in seen:
            continue
        seen.add(key)
        elections.append(election)
        pending_type = ""

    return elections


def parse_ballot_count_history(rows: list[dict]) -> list[OrBallotReturn]:
    parsed: list[OrBallotReturn] = []
    for row in rows:
        election = _clean_text(str(row.get("election") or ""))
        election_date = _parse_election_label_date(election)
        count_date = _parse_iso_date(str(row.get("date") or ""))
        if not election or not election_date or not count_date:
            continue
        parsed.append(
            OrBallotReturn(
                election=election,
                election_date=election_date,
                count_date=count_date,
                daily_ballots_returned=_safe_int(row.get("number_of_ballots_returned") or ""),
                cumulative_ballots_returned=_safe_int(row.get("cumulative_number_of_ballots") or ""),
                daily_return_pct_of_total_ballots=_safe_float(row.get("daily_return_as_of_total") or ""),
                daily_return_pct_of_total_return=_safe_float(row.get("daily_return_as_of_total_1") or ""),
                cumulative_return_pct_of_total_ballots=_safe_float(row.get("cumulative_return_as_of_total") or ""),
                cumulative_return_pct_of_total_return=_safe_float(row.get("cumulative_return_as_of_total_1") or ""),
            )
        )
    return parsed


def latest_ballot_returns_by_election(records: list[OrBallotReturn]) -> dict[datetime.date, OrBallotReturn]:
    latest: dict[datetime.date, OrBallotReturn] = {}
    for record in records:
        existing = latest.get(record.election_date)
        if existing is None or record.count_date > existing.count_date:
            latest[record.election_date] = record
    return latest


def ballot_return_payload(record: OrBallotReturn, source_url: str = "") -> dict:
    return {
        "election": record.election,
        "election_date": record.election_date.isoformat(),
        "count_date": record.count_date.isoformat(),
        "daily_ballots_returned": record.daily_ballots_returned,
        "cumulative_ballots_returned": record.cumulative_ballots_returned,
        "daily_return_pct_of_total_ballots": record.daily_return_pct_of_total_ballots,
        "daily_return_pct_of_total_return": record.daily_return_pct_of_total_return,
        "cumulative_return_pct_of_total_ballots": record.cumulative_return_pct_of_total_ballots,
        "cumulative_return_pct_of_total_return": record.cumulative_return_pct_of_total_return,
        "source_url": source_url,
    }


def parse_candidate_filings(html: str) -> list[OrCandidateFiling]:
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table", id="cfSearchResults")
    if table is None:
        return []
    rows: list[OrCandidateFiling] = []
    for row in table.find_all("tr")[1:]:
        cells = [_clean_text(cell.get_text(" ", strip=True)) for cell in row.find_all("td")]
        if len(cells) < 7:
            continue
        rows.append(
            OrCandidateFiling(
                ballot_name=cells[0],
                party=cells[1],
                office=cells[2],
                election=cells[3],
                filing_method=cells[4],
                filing_date=cells[5],
                qualified=cells[6],
            )
        )
    return rows


def parse_local_measures(html: str) -> list[OrLocalMeasure]:
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table", id="measSearchResults")
    if table is None:
        return []
    rows: list[OrLocalMeasure] = []
    for row in table.find_all("tr")[1:]:
        cells = [_clean_text(cell.get_text(" ", strip=True)) for cell in row.find_all("td")]
        if len(cells) < 4:
            continue
        rows.append(
            OrLocalMeasure(
                measure_number=cells[0],
                election=cells[1],
                county=cells[2],
                ballot_title_caption=cells[3],
            )
        )
    return rows


def _election_from_line(line: str, source_url: str = "") -> OrElectionInfo | None:
    date_match = _DATE_RE.search(line)
    if not date_match:
        return None

    election_type = _election_type_from_line(line)
    if not election_type:
        return None

    date_text = date_match.group(0)
    name = _clean_election_name(line, date_text, election_type)
    return OrElectionInfo(
        name=name,
        election_date=date_text,
        election_type=election_type,
        source_url=source_url,
    )


def _election_type_from_line(line: str) -> str:
    lower = line.lower()
    if "election" not in lower:
        return ""
    if "primary" in lower:
        return "primary"
    if "general" in lower:
        return "general"
    if "special" in lower:
        return "special"
    return ""


def _heading_election_type_from_line(line: str) -> str:
    lower = line.lower().strip(" .:-")
    if lower in {"primary election", "general election", "special election"}:
        return _election_type_from_line(line)
    return ""


def _election_info_from_date(date_text: str, election_type: str, source_url: str = "") -> OrElectionInfo:
    label = election_type.replace("_", " ").title()
    return OrElectionInfo(
        name=f"Oregon {label} Election",
        election_date=date_text,
        election_type=election_type,
        source_url=source_url,
    )


def _clean_election_name(line: str, date_text: str, election_type: str) -> str:
    cleaned = _clean_text(line.replace(date_text, ""))
    cleaned = re.sub(r"^[\-–:|\s]+|[\-–:|\s]+$", "", cleaned)
    if "election" not in cleaned.lower():
        label = election_type.replace("_", " ").title()
        cleaned = f"Oregon {label} Election"
    if not cleaned.lower().startswith("oregon"):
        cleaned = f"Oregon {cleaned}"
    return cleaned


def _parse_election_label_date(value: str) -> datetime.date | None:
    match = re.search(
        r"\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}\b",
        value,
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    try:
        return datetime.datetime.strptime(match.group(0), "%B %d, %Y").date()
    except ValueError:
        return None


def _parse_iso_date(value: str) -> datetime.date | None:
    if not value:
        return None
    try:
        return datetime.date.fromisoformat(value[:10])
    except ValueError:
        return None


def parse_open_offices_text(text: str) -> list[OrOpenOffice]:
    offices: list[OrOpenOffice] = []
    seen: set[tuple[str, str]] = set()
    section = ""

    for raw_line in text.splitlines():
        line = _clean_text(raw_line)
        if not line:
            continue

        new_section = _open_office_section(line)
        if new_section is not None:
            section = new_section

        for office in _offices_from_line(line, section):
            key = (office.office_code, office.district)
            if key in seen:
                continue
            seen.add(key)
            offices.append(office)

    return offices


_DATE_RE = re.compile(
    r"\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}\b",
    flags=re.IGNORECASE,
)


_PARTY_LABELS = {"democrat", "republican", "independent", "libertarian", "nonpartisan"}


def _parse_result_pdf_page(page_text: str) -> list[OrResultRecord]:
    lines = [_clean_text(line) for line in page_text.splitlines()]
    lines = [line for line in lines if line]
    if not lines:
        return []

    try:
        county_index = next(idx for idx, line in enumerate(lines) if line.startswith("County"))
    except StopIteration:
        return []

    total_line = next((line for line in lines[county_index + 1:] if line.startswith("Total ")), "")
    if not total_line:
        return []

    preamble = [
        line
        for line in lines[:county_index]
        if "Abstract of Votes" not in line and not line.startswith("* Nominee") and not line.startswith("** Elected")
    ]
    party_index = next((idx for idx, line in enumerate(preamble) if _is_party_pdf_line(line)), None)
    if party_index is None:
        party = ""
        if preamble and preamble[0].lower().startswith("measure "):
            office_lines = preamble[:1]
            choice_lines = []
        else:
            office_lines = preamble[:-1]
            choice_lines = preamble[-1:]
    else:
        party = _clean_party_pdf_line(preamble[party_index])
        office_lines = preamble[:party_index]
        choice_lines = preamble[party_index + 1:]

    office_title = _normalize_result_pdf_office_title(office_lines)
    if not office_title:
        return []

    choices = _pdf_choices_from_header(choice_lines, lines[county_index])
    votes = [_safe_int(value) for value in re.findall(r"\d[\d,]*", total_line)]
    if not choices or not votes:
        return []

    records: list[OrResultRecord] = []
    for choice, vote_count in zip(choices, votes, strict=False):
        records.append(
            OrResultRecord(
                office_title=office_title,
                choice=choice,
                vote_count=vote_count,
                party=party,
                source_file="official-results.pdf",
            )
        )
    return records


def _is_party_pdf_line(line: str) -> bool:
    lowered = line.lower().replace("(cont.)", "").strip()
    return lowered in _PARTY_LABELS


def _clean_party_pdf_line(line: str) -> str:
    return _clean_text(line.replace("(cont.)", "")).title()


def _normalize_result_pdf_office_title(lines: list[str]) -> str:
    office_lines = [line for line in lines if line]
    if not office_lines:
        return ""
    office = " ".join(office_lines)
    district = _result_pdf_district(office)
    lowered = office.lower()
    if "us senator" in lowered or "u.s. senator" in lowered:
        return "U.S. Senator"
    if "us representative" in lowered or "u.s. representative" in lowered:
        return f"U.S. Representative, District {district}" if district else "U.S. Representative"
    if "state senator" in lowered or "state senate" in lowered:
        return f"Oregon State Senator, District {district}" if district else "Oregon State Senator"
    if "state representative" in lowered or "state house" in lowered:
        return f"Oregon State Representative, District {district}" if district else "Oregon State Representative"
    return office


def _result_pdf_district(value: str) -> str:
    match = re.search(r"\b(\d+)(?:st|nd|rd|th)?\s+District\b", value, flags=re.IGNORECASE)
    return str(int(match.group(1))) if match else ""


def _pdf_choices_from_header(choice_lines: list[str], county_line: str) -> list[str]:
    if choice_lines:
        raw_choices = choice_lines[0]
    else:
        raw_choices = county_line.removeprefix("County").strip()
    return [_clean_pdf_choice(choice) for choice in raw_choices.split() if _clean_pdf_choice(choice)]


def _clean_pdf_choice(value: str) -> str:
    return value.strip().strip("*").strip()


def _offices_from_line(line: str, section: str = "") -> list[OrOpenOffice]:
    norm = line.lower()
    offices: list[OrOpenOffice] = []

    if re.search(r"\bu\.?s\.?\s+senator\b", norm) or "united states senator" in norm:
        offices.append(OrOpenOffice(office_title="U.S. Senator", office_code="us_senate"))

    if re.search(r"\bu\.?s\.?\s+representative\b", norm) or "united states representative" in norm:
        for district in _extract_districts(line):
            offices.append(
                OrOpenOffice(
                    office_title=f"U.S. Representative, District {district}",
                    office_code="us_house",
                    district=district,
                    ocd_division_id=f"ocd-division/country:us/state:or/cd:{district}",
                    geography_scope="district",
                )
            )

    if section == "us_house" and _district_line_number(line):
        district = _district_line_number(line)
        offices.append(
            OrOpenOffice(
                office_title=f"U.S. Representative, District {district}",
                office_code="us_house",
                district=district,
                ocd_division_id=f"ocd-division/country:us/state:or/cd:{district}",
                geography_scope="district",
            )
        )

    if re.fullmatch(r".*\bgovernor\b.*", norm) and "lieutenant" not in norm:
        offices.append(OrOpenOffice(office_title="Governor", office_code="governor"))

    if "state senator" in norm or "state senate" in norm:
        for district in _extract_districts(line):
            offices.append(
                OrOpenOffice(
                    office_title=f"Oregon State Senate, District {district}",
                    office_code="state_senate",
                    district=district,
                    ocd_division_id=f"ocd-division/country:us/state:or/sldu:{district}",
                    geography_scope="district",
                )
            )

    if section == "state_senate" and _district_line_number(line):
        district = _district_line_number(line)
        offices.append(
            OrOpenOffice(
                office_title=f"Oregon State Senate, District {district}",
                office_code="state_senate",
                district=district,
                ocd_division_id=f"ocd-division/country:us/state:or/sldu:{district}",
                geography_scope="district",
            )
        )

    if "state representative" in norm or "state house" in norm:
        for district in _extract_districts(line):
            offices.append(
                OrOpenOffice(
                    office_title=f"Oregon State Representative, District {district}",
                    office_code="state_house",
                    district=district,
                    ocd_division_id=f"ocd-division/country:us/state:or/sldl:{district}",
                    geography_scope="district",
                )
            )

    if section == "state_house" and _district_line_number(line):
        district = _district_line_number(line)
        offices.append(
            OrOpenOffice(
                office_title=f"Oregon State Representative, District {district}",
                office_code="state_house",
                district=district,
                ocd_division_id=f"ocd-division/country:us/state:or/sldl:{district}",
                geography_scope="district",
            )
        )

    return offices


def _open_office_section(line: str) -> str | None:
    norm = line.strip().lower()
    if norm in {"us representative", "u.s. representative", "united states representative"}:
        return "us_house"
    if norm in {"state senate", "state senator"}:
        return "state_senate"
    if norm in {"state representative", "state house"}:
        return "state_house"
    if norm in {"governor", "us senator", "u.s. senator", "nonpartisan offices", "judge of the circuit court"}:
        return ""
    return None


def _district_line_number(line: str) -> str:
    match = re.match(r"^(\d+)(?:st|nd|rd|th)\s+District\b", line, flags=re.IGNORECASE)
    return str(int(match.group(1))) if match else ""


def _extract_districts(line: str) -> list[str]:
    match = re.search(r"districts?\s+([0-9,\s\-–]+)", line, flags=re.IGNORECASE)
    if not match:
        return []

    districts: list[str] = []
    for part in re.split(r",|\band\b", match.group(1), flags=re.IGNORECASE):
        part = part.strip()
        if not part:
            continue
        range_match = re.fullmatch(r"(\d+)\s*[-–]\s*(\d+)", part)
        if range_match:
            start, end = int(range_match.group(1)), int(range_match.group(2))
            districts.extend(str(n) for n in range(start, end + 1))
        elif part.isdigit():
            districts.append(str(int(part)))
    return districts


def _parse_zip(content: bytes) -> list[OrResultRecord]:
    records: list[OrResultRecord] = []
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        for name in archive.namelist():
            suffix = PurePosixPath(name).suffix.lower()
            if suffix not in {".csv", ".txt", ".tsv"}:
                if suffix == ".xlsx":
                    records.extend(_parse_xlsx(archive.read(name), source_file=name))
                continue
            records.extend(_parse_delimited(archive.read(name), source_file=name))
    return records


def _parse_xlsx(content: bytes, source_file: str = "") -> list[OrResultRecord]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    records: list[OrResultRecord] = []
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            header_index = _find_header_row(rows)
            if header_index is None:
                continue
            headers = [_clean_text(str(value or "")) for value in rows[header_index]]
            for values in rows[header_index + 1:]:
                row = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values))) if headers[idx]}
                office = _first(row, "contest", "contest name", "office", "office title", "race")
                choice = _first(row, "candidate", "candidate name", "choice", "option", "selection")
                votes = _safe_int(_first(row, "votes", "vote count", "total votes", "total"))
                if not office or not choice:
                    continue
                records.append(
                    OrResultRecord(
                        office_title=office,
                        choice=choice,
                        vote_count=votes,
                        vote_pct=_safe_float(_first(row, "vote pct", "vote percent", "percent", "percentage")),
                        party=_first(row, "party", "party name", "candidate party"),
                        jurisdiction=_first(row, "county", "jurisdiction", "reporting unit"),
                        source_file=source_file or sheet.title,
                    )
                )
    finally:
        workbook.close()
    return records


def _find_header_row(rows: list[tuple]) -> int | None:
    for idx, values in enumerate(rows[:20]):
        normalized = {_normalize_key(str(value or "")) for value in values}
        has_office = bool(normalized & {"contest", "contest name", "office", "office title", "race"})
        has_choice = bool(normalized & {"candidate", "candidate name", "choice", "option", "selection"})
        has_votes = bool(normalized & {"votes", "vote count", "total votes", "total"})
        if has_office and has_choice and has_votes:
            return idx
    return None


def _parse_delimited(content: bytes, source_file: str = "") -> list[OrResultRecord]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if not reader.fieldnames:
        return []

    records: list[OrResultRecord] = []
    for row in reader:
        office = _first(row, "contest", "contest name", "office", "office title", "race")
        choice = _first(row, "candidate", "candidate name", "choice", "option", "selection")
        votes = _safe_int(_first(row, "votes", "vote count", "total votes", "total"))
        if not office or not choice:
            continue
        records.append(
            OrResultRecord(
                office_title=office,
                choice=choice,
                vote_count=votes,
                vote_pct=_safe_float(_first(row, "vote pct", "vote percent", "percent", "percentage")),
                party=_first(row, "party", "party name", "candidate party"),
                jurisdiction=_first(row, "county", "jurisdiction", "reporting unit"),
                source_file=source_file,
            )
        )
    return records


def _first(row: dict, *names: str) -> str:
    normalized = {_normalize_key(key): value for key, value in row.items()}
    for name in names:
        value = normalized.get(_normalize_key(name))
        if value is not None and str(value).strip():
            return _clean_text(str(value))
    return ""


def _normalize_key(value: str) -> str:
    return " ".join((value or "").strip().lower().replace("_", " ").split())


def _clean_text(value: str) -> str:
    return " ".join("".join(ch for ch in value if ch >= " " or ch == "\t").split())


def _safe_int(value: str) -> int:
    try:
        return int(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0


def _safe_float(value: str) -> float | None:
    try:
        cleaned = str(value).replace("%", "").strip()
        return float(cleaned) if cleaned else None
    except (TypeError, ValueError):
        return None
