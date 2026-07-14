from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from urllib.parse import urljoin, urlparse

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .client import ELECTION_CALENDAR_URL, RESULTS_INDEX_URL


@dataclass(frozen=True)
class TnElectionRow:
    name: str
    election_date: date
    county: str
    jurisdiction: str
    source_url: str
    is_statewide: bool


@dataclass(frozen=True)
class TnCandidateWorkbookLink:
    office_group: str
    filename: str
    url: str


@dataclass(frozen=True)
class TnCandidateRecord:
    office: str
    district: str
    candidate_name: str
    party: str
    status: str
    source_url: str
    source_row: int


@dataclass(frozen=True)
class TnResultLink:
    election_date: date | None
    label: str
    url: str
    file_type: str
    result_level: str
    source_version: str


@dataclass(frozen=True)
class TnResultRecord:
    county: str
    precinct: str
    office_title: str
    candidate_name: str
    party: str
    vote_count: int
    source_url: str


_MONTH_DATE_RE = re.compile(
    r"\b(?:January|February|March|April|May|June|July|August|September|October|November|December)"
    r"\s+\d{1,2},\s+\d{4}\b"
)
_SLASH_DATE_RE = re.compile(r"\b\d{1,2}/\d{1,2}/\d{4}\b")
_TN_SOS_PAGE_HOST = "sos.tn.gov"
_TN_SOS_DOCUMENT_HOST = "sos-prod.tnsosgovfiles.com"
_TN_SOS_DOCUMENT_PATH = "/s3fs-public/document/"
_POSITIVE_CANDIDATE_STATUSES = {
    "active",
    "active candidate",
    "nominee",
    "nominated",
    "nominated candidate",
    "qualified",
    "qualified candidate",
}


def document_checksum(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def parse_calendar(html: str) -> list[TnElectionRow]:
    soup = BeautifulSoup(html, "lxml")
    rows: list[TnElectionRow] = []

    for heading in soup.find_all(["h2", "h3", "h4", "h5"]):
        name = _clean_text(heading.get_text(" ", strip=True))
        election_date = _parse_date(name)
        if election_date is None or "election" not in name.lower():
            continue
        rows.append(
            TnElectionRow(
                name=name,
                election_date=election_date,
                county="Tennessee",
                jurisdiction="",
                source_url=ELECTION_CALENDAR_URL,
                is_statewide=True,
            )
        )

    for table in soup.find_all("table"):
        headers = _table_headers(table)
        if not {"county", "jurisdiction", "date"}.issubset(headers.values()):
            continue
        for table_row in table.find_all("tr"):
            cells = [_clean_text(cell.get_text(" ", strip=True)) for cell in table_row.find_all("td")]
            if len(cells) < 3:
                continue
            values = {header: cells[index] for index, header in headers.items() if index < len(cells)}
            election_date = _parse_date(values.get("date", ""))
            county = values.get("county", "")
            jurisdiction = values.get("jurisdiction", "")
            if not election_date or not county or not jurisdiction:
                continue
            rows.append(
                TnElectionRow(
                    name=f"{jurisdiction} Election",
                    election_date=election_date,
                    county=county,
                    jurisdiction=jurisdiction,
                    source_url=ELECTION_CALENDAR_URL,
                    is_statewide=False,
                )
            )
    return rows


def parse_candidate_workbook_links(html: str) -> list[TnCandidateWorkbookLink]:
    soup = BeautifulSoup(html, "lxml")
    links: list[TnCandidateWorkbookLink] = []
    for anchor in soup.find_all("a", href=True):
        url = urljoin("https://sos.tn.gov", anchor["href"])
        if _is_official_tn_sos_url(url) and urlparse(url).path.lower().endswith(".xlsx"):
            container = anchor.find_parent("li")
            office_group = _clean_text(container.get_text(" ", strip=True) if container else anchor.get_text(" ", strip=True))
            office_group = re.split(r"\s*:\s*", office_group, maxsplit=1)[0]
            links.append(
                TnCandidateWorkbookLink(
                    office_group=office_group,
                    filename=urlparse(url).path.rsplit("/", 1)[-1],
                    url=url,
                )
            )
    return links


def parse_candidate_workbook(content: bytes, source_url: str) -> list[TnCandidateRecord]:
    records: list[TnCandidateRecord] = []
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            header_index = _find_header_row(rows, {"office"}, {"candidate", "candidate name"})
            if header_index is None:
                continue
            headers = [_normalize_key(value) for value in rows[header_index]]
            has_status_field = any(header in {"status", "candidate status"} for header in headers)
            for source_row, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
                row = _row_values(headers, values)
                office = _first(row, "office", "office title", "contest")
                candidate_name = _first(row, "candidate", "candidate name", "name")
                if not office or not candidate_name:
                    continue
                status = _first(row, "status", "candidate status")
                if has_status_field and not _is_positive_candidate_status(status):
                    continue
                records.append(
                    TnCandidateRecord(
                        office=office,
                        district=_first(row, "district"),
                        candidate_name=candidate_name,
                        party=_first(row, "party", "party name"),
                        status=status,
                        source_url=source_url,
                        source_row=source_row,
                    )
                )
    finally:
        workbook.close()
    return records


def parse_results_index(html: str) -> list[TnResultLink]:
    soup = BeautifulSoup(html, "lxml")
    links: list[TnResultLink] = []
    for anchor in soup.find_all("a", href=True):
        url = urljoin(RESULTS_INDEX_URL, anchor["href"])
        path = urlparse(url).path
        extension = path.rsplit(".", 1)[-1].lower() if "." in path.rsplit("/", 1)[-1] else ""
        if not _is_official_tn_sos_url(url) or extension not in {"xlsx", "xls", "pdf", "csv"}:
            continue
        label = _clean_text(anchor.get_text(" ", strip=True))
        context = _clean_text(anchor.find_parent("li").get_text(" ", strip=True) if anchor.find_parent("li") else label)
        links.append(
            TnResultLink(
                election_date=_date_from_ancestors(anchor),
                label=label,
                url=url,
                file_type=extension,
                result_level=_result_level(f"{label} {context} {path}"),
                source_version=path.rsplit("/", 1)[-1],
            )
        )
    return links


def parse_precinct_xlsx(content: bytes, source_url: str) -> list[TnResultRecord]:
    records: list[TnResultRecord] = []
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            header_index = _find_header_row(rows, {"office", "office title", "contest"}, {"candidate", "candidate name"}, {"votes", "vote count", "total votes"})
            if header_index is None:
                continue
            headers = [_normalize_key(value) for value in rows[header_index]]
            for values in rows[header_index + 1 :]:
                row = _row_values(headers, values)
                office = _first(row, "office", "office title", "contest", "race")
                candidate = _first(row, "candidate", "candidate name", "choice")
                vote_count = _parse_int(_first(row, "votes", "vote count", "total votes", "total"))
                if not office or not candidate or vote_count is None:
                    continue
                records.append(
                    TnResultRecord(
                        county=_first(row, "county"),
                        precinct=_first(row, "precinct", "precinct name", "reporting unit"),
                        office_title=office,
                        candidate_name=candidate,
                        party=_first(row, "party", "party name", "candidate party"),
                        vote_count=vote_count,
                        source_url=source_url,
                    )
                )
    finally:
        workbook.close()
    return records


def _table_headers(table) -> dict[int, str]:
    header_cells = table.find_all("th")
    headers: dict[int, str] = {}
    for index, cell in enumerate(header_cells):
        normalized = _normalize_key(cell.get_text(" ", strip=True))
        if normalized == "election":
            normalized = "jurisdiction"
        if normalized in {"county", "jurisdiction", "date"}:
            headers[index] = normalized
    return headers


def _find_header_row(rows: list[tuple], *required_groups: set[str]) -> int | None:
    for index, values in enumerate(rows[:20]):
        headers = {_normalize_key(value) for value in values}
        if all(headers & group for group in required_groups):
            return index
    return None


def _row_values(headers: list[str], values: tuple) -> dict[str, object]:
    return {header: values[index] for index, header in enumerate(headers) if header and index < len(values)}


def _first(row: dict[str, object], *names: str) -> str:
    for name in names:
        value = row.get(_normalize_key(name))
        if value is not None and str(value).strip():
            return _clean_text(str(value))
    return ""


def _date_from_ancestors(anchor) -> date | None:
    for parent in anchor.parents:
        if getattr(parent, "name", None) != "li":
            continue
        parsed = _parse_date(_clean_text(parent.get_text(" ", strip=True)))
        if parsed:
            return parsed
    return None


def _parse_date(value: str) -> date | None:
    match = _MONTH_DATE_RE.search(value) or _SLASH_DATE_RE.search(value)
    if not match:
        return None
    for format_string in ("%B %d, %Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(match.group(0), format_string).date()
        except ValueError:
            continue
    return None


def _parse_int(value: str) -> int | None:
    try:
        return int(value.replace(",", ""))
    except (AttributeError, TypeError, ValueError):
        return None


def _result_level(value: str) -> str:
    normalized = _normalize_key(value)
    if "precinct" in normalized:
        return "precinct"
    if "county" in normalized:
        return "county"
    if "office" in normalized:
        return "office"
    return "unknown"


def _normalize_key(value: object) -> str:
    return " ".join(str(value or "").lower().replace("_", " ").split())


def _is_official_tn_sos_url(url: str) -> bool:
    parsed = urlparse(url)
    if parsed.scheme != "https" or parsed.hostname is None:
        return False
    hostname = parsed.hostname.lower()
    if hostname == _TN_SOS_PAGE_HOST:
        return True
    return hostname == _TN_SOS_DOCUMENT_HOST and parsed.path.lower().startswith(_TN_SOS_DOCUMENT_PATH)


def _is_positive_candidate_status(status: str) -> bool:
    normalized = " ".join(re.sub(r"[^a-z0-9]+", " ", status.lower()).split())
    return normalized in _POSITIVE_CANDIDATE_STATUSES


def _clean_text(value: str) -> str:
    return " ".join(value.replace("\u200b", "").split())
